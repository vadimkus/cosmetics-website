#!/usr/bin/env python3
"""Add CERABARRIER 200ml + 600ml clinic lines to July 2026 price list workbook."""
from __future__ import annotations

import csv
import io
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import range_boundaries
from PIL import Image

XLSX = Path(
    '/Users/vadimkus/Desktop/Drive/Genosys/Price_list/2026/July/'
    'GENOSYS_UAE_PriceList_Clinics_2026.xlsx'
)
CSV_OUT = Path(__file__).resolve().parents[1] / 'docs/GENOSYS_UAE_PriceList_Clinics_2026_normalized.csv'
IMG_200 = Path(__file__).resolve().parents[1] / 'public/images/cera/cerabar_200ml.jpeg'
IMG_600 = Path(__file__).resolve().parents[1] / 'public/images/cera/cerabar_600ml.jpeg'

PRODUCT = 'CERABARRIER BIOME GEL CLEANSER'
INSERT_AT = 73  # after SNOW O₂ block (rows 69-72)
INSERT_COUNT = 4


def copy_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def collect_merges(ws):
    out = []
    for m in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(m))
        out.append((min_row, max_row, min_col, max_col))
    return out


def unmerge_all(ws):
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))


def apply_merges(ws, merges):
    for min_row, max_row, min_col, max_col in merges:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def shift_merges(merges, insert_at: int, count: int):
    shifted = []
    for min_row, max_row, min_col, max_col in merges:
        if min_row >= insert_at:
            min_row += count
            max_row += count
        elif max_row >= insert_at:
            max_row += count
        shifted.append((min_row, max_row, min_col, max_col))
    return shifted


def shift_images(ws, insert_at: int, count: int):
    for im in getattr(ws, '_images', []):
        try:
            if im.anchor._from.row + 1 >= insert_at:
                im.anchor._from.row += count
                if hasattr(im.anchor, '_to') and im.anchor._to is not None:
                    im.anchor._to.row += count
        except Exception:
            pass


def add_image(ws, path: Path, row: int, col: int = 3):
    data = path.read_bytes()
    try:
        img = XLImage(io.BytesIO(data))
    except Exception:
        buf = io.BytesIO()
        Image.open(io.BytesIO(data)).save(buf, format='PNG')
        buf.seek(0)
        img = XLImage(buf)
    img.width = 90
    img.height = 90
    col_letter = chr(ord('A') + col - 1)
    img.anchor = f'{col_letter}{row}'
    ws.add_image(img)


def extract_items(ws):
    items = []
    cat = None
    last_product = None
    for r in range(1, ws.max_row + 1):
        c3 = ws.cell(r, 3).value
        c4, c5, c6, c7, c8, c9 = [ws.cell(r, c).value for c in range(4, 10)]

        if c3 and str(c3).strip().startswith('GENOSYS') and not c4 and not c5:
            cat = str(c3).strip()
        if c4 and not c5 and not c6 and not c7 and not c8 and not c9:
            if 'GENOSYS' in str(c4) or 'Genosys' in str(c4):
                cat = str(c4).strip()

        price_raw = c9 if c9 is not None else c8
        if price_raw is None:
            continue
        pr = str(price_raw).strip()
        if pr in ('', 'Price                  AED'):
            continue
        if pr.upper() == 'N/A':
            price = 'N/A'
        else:
            try:
                price = float(price_raw)
            except (TypeError, ValueError):
                continue

        product = None
        description = ''
        qty = ''
        unit = ''
        if c5 and not str(c5).strip().endswith('use') and c6 and c8 is not None and c9 is None and c4:
            # page 1 layout: E=product, F=desc, G=qty, H=unit, I=price
            product = str(c4).strip() if c4 else last_product
            description = str(c5).strip()
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''
        elif c5:
            # page 2 layout: D=product (optional), E=desc, F=qty, G=unit, H=price
            if c4:
                product = str(c4).strip()
                last_product = product
            else:
                product = last_product
            description = str(c5).strip()
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''
        elif c4:
            product = str(c4).strip()
            last_product = product
            description = str(c5).strip() if c5 else ''
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''

        if not product:
            continue
        last_product = product
        items.append(
            {
                'row': r,
                'category': cat or '',
                'product': product,
                'description': description,
                'quantity_or_spec': qty,
                'unit': unit,
                'price_aed': int(price) if isinstance(price, float) and price == int(price) else price,
            }
        )
    return items


def main():
    wb = load_workbook(XLSX)
    ws = wb.active
    merges = collect_merges(ws)
    unmerge_all(ws)
    shift_images(ws, INSERT_AT, INSERT_COUNT)
    ws.insert_rows(INSERT_AT, INSERT_COUNT)
    merges = shift_merges(merges, INSERT_AT, INSERT_COUNT)

    r_home, r_home2 = INSERT_AT, INSERT_AT + 1
    r_pro, r_pro2 = INSERT_AT + 2, INSERT_AT + 3

    # Mirror SNOW O₂ merge pattern (cols C-H = 3-8)
    merges.extend(
        [
            (r_home, r_home2, 3, 3),
            (r_home, r_home2, 4, 4),
            (r_home, r_home2, 5, 5),
            (r_home, r_home2, 6, 6),
            (r_home, r_home2, 7, 7),
            (r_home, r_home2, 8, 8),
            (r_pro, r_pro2, 3, 3),
            (r_pro, r_pro2, 5, 5),
            (r_pro, r_pro2, 6, 6),
            (r_pro, r_pro2, 7, 7),
            (r_pro, r_pro2, 8, 8),
            (r_home, r_pro2, 4, 4),  # product name spans all 4 rows
        ]
    )
    apply_merges(ws, merges)

    template = 69
    for r in (r_home, r_pro):
        for c in range(3, 9):
            copy_style(ws.cell(template, c), ws.cell(r, c))

    ws.cell(r_home, 4, PRODUCT)
    ws.cell(r_home, 5, 'Daily biome gel cleanser / Personal use')
    ws.cell(r_home, 6, '200ml')
    ws.cell(r_home, 7, 'Pcs')
    ws.cell(r_home, 8, 190)

    ws.cell(r_pro, 5, 'Professional biome gel cleanser')
    ws.cell(r_pro, 6, '600ml')
    ws.cell(r_pro, 7, 'Pcs')
    ws.cell(r_pro, 8, 310)

    if IMG_200.exists():
        add_image(ws, IMG_200, r_home)
    if IMG_600.exists():
        add_image(ws, IMG_600, r_pro)

    wb.save(XLSX)

    wb2 = load_workbook(XLSX, data_only=True)
    items = extract_items(wb2.active)
    cera = [x for x in items if 'CERABARRIER' in x['product'].upper()]
    if len(cera) != 2:
        raise RuntimeError(f'Expected 2 CERABARRIER lines, got {cera}')

    with CSV_OUT.open('w', newline='') as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                'row',
                'category',
                'product',
                'description',
                'quantity_or_spec',
                'unit',
                'price_aed',
            ],
        )
        w.writeheader()
        w.writerows(items)

    print(f'OK: {len(items)} priced lines ({len(cera)} CERABARRIER)')
    for x in cera:
        print(
            f"  R{x['row']}: {x['product']} | {x['description']} | "
            f"{x['quantity_or_spec']} | {x['unit']} | {x['price_aed']} AED"
        )


if __name__ == '__main__':
    main()
