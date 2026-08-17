#!/usr/bin/env python3
"""Update July 2026 clinic price list: Bio Meso PDRN + roller Standard/Narrow merge.

Safe for heavily merged GENOSYS workbook: unmerge → edit → insert → shift merges → remerge.
"""
from __future__ import annotations

import csv
import io
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import range_boundaries
from PIL import Image

SRC = Path(
    '/Users/vadimkus/Desktop/Drive/Genosys/Price_list/2026/July/'
    'GENOSYS_UAE_PriceList_Clinics_2026.xlsx'
)
BACKUP = SRC.with_name(
    'GENOSYS_UAE_PriceList_Clinics_2026.before_20260713_bio_meso_roller_merge.xlsx'
)
CSV_OUT = Path(__file__).resolve().parents[1] / 'docs/GENOSYS_UAE_PriceList_Clinics_2026_normalized.csv'
IMG_6000 = Path(__file__).resolve().parents[1] / 'public/images/6000/main.jpg'
IMG_5000 = Path(__file__).resolve().parents[1] / 'public/images/meso_5000/main.jpg'


def copy_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def collect_merges(ws):
    out = []
    for m in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(m))
        out.append((min_row, max_row, min_col, max_col))
    return out


def unmerge_all(ws):
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))


def apply_merges(ws, merges):
    for min_row, max_row, min_col, max_col in merges:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def shift_merges(merges, insert_at: int, count: int):
    shifted = []
    for min_row, max_row, min_col, max_col in merges:
        if min_row >= insert_at:
            min_row += count
            max_row += count
        elif max_row >= insert_at:
            max_row += count
        shifted.append((min_row, max_row, min_col, max_col))
    return shifted


def extract_items(ws):
    items = []
    cat = None
    last = None
    for r in range(1, ws.max_row + 1):
        d, e, f, g, h, i = [ws.cell(r, c).value for c in range(4, 10)]
        if d and not e and not f and not g and not h and (i is None or str(i).strip() == ''):
            cat = str(d).strip()
            continue
        if i is None:
            continue
        pr = str(i).strip()
        if pr in ('', 'Price                  AED'):
            continue
        if pr.upper() == 'N/A':
            price = 'N/A'
        else:
            try:
                price = float(i)
            except (TypeError, ValueError):
                continue
        prod = str(e).strip() if e else last
        if not prod:
            continue
        last = prod
        items.append(
            {
                'row': r,
                'category': cat or '',
                'product': prod,
                'description': str(f).strip() if f else '',
                'quantity_or_spec': str(g).strip() if g else '',
                'unit': str(h).strip() if h else '',
                'price_aed': int(price) if isinstance(price, float) and price == int(price) else price,
            }
        )
    return items


def get_image_bytes(im):
    try:
        data = im._data()
        return data() if callable(data) else data
    except Exception:
        try:
            return im.ref.getvalue()
        except Exception:
            return None


def item_key(it):
    return (
        it['category'],
        ' '.join(it['product'].split()),
        ' '.join(it['description'].split()),
        ' '.join(it['quantity_or_spec'].split()),
        (it['unit'] or '').strip(),
    )


ROLLER_IMAGE_FROM = {
    'Detachable Manual Roller': 'Standard Detachable Manual Roller',
    'Manual Roller': 'Standard Manual Roller',
}


def reanchor_images(ws, items):
    """Re-map column D photos from pre-edit backup onto current product rows."""
    wb_b = load_workbook(BACKUP)
    ws_b = wb_b.active
    backup_items = extract_items(ws_b)
    backup_by_row = {x['row']: x for x in backup_items}

    anchor_bytes = {}
    for im in getattr(ws_b, '_images', []):
        try:
            r0, c0 = im.anchor._from.row, im.anchor._from.col
        except Exception:
            continue
        if c0 != 3:
            continue
        row = r0 + 1
        data = get_image_bytes(im)
        if not data:
            continue
        if row not in anchor_bytes or len(data) > len(anchor_bytes[row]):
            anchor_bytes[row] = data

    key_to_image = {}
    orphan_rows = sorted(set(anchor_bytes) - set(backup_by_row))
    for it in backup_items:
        row = it['row']
        if row in anchor_bytes:
            key_to_image[item_key(it)] = anchor_bytes[row]
    for it in backup_items:
        k = item_key(it)
        if k in key_to_image:
            continue
        p = it['row']
        cands = [a for a in orphan_rows if abs(a - p) <= 3]
        if cands:
            cands.sort(key=lambda a: (abs(a - p), 0 if a > p else 1))
            key_to_image[k] = anchor_bytes[cands[0]]

    backup_name_to_row = {x['product']: x['row'] for x in backup_items}
    bio_files = {
        'BIO-MESO PDRN Expert Ampoule 60000': IMG_6000,
        'BIO-MESO PDRN Homecare Ampoule 5000': IMG_5000,
    }

    row_image = {}
    for cur in items:
        row = int(cur['row'])
        data = None
        if cur['product'] in bio_files and bio_files[cur['product']].exists():
            data = bio_files[cur['product']].read_bytes()
        elif item_key(cur) in key_to_image:
            data = key_to_image[item_key(cur)]
        elif cur['product'] in ROLLER_IMAGE_FROM:
            src_name = ROLLER_IMAGE_FROM[cur['product']]
            src_row = backup_name_to_row.get(src_name)
            if src_row and src_row in anchor_bytes:
                data = anchor_bytes[src_row]
        if not data:
            raise RuntimeError(f'No image mapped for row {row}: {cur["product"]}')
        row_image[row] = data

    ws._images = []
    for row, data in sorted(row_image.items()):
        try:
            img = XLImage(io.BytesIO(data))
        except Exception:
            pil = Image.open(io.BytesIO(data))
            buf = io.BytesIO()
            pil.save(buf, format='PNG')
            buf.seek(0)
            img = XLImage(buf)
        img.width = 90
        img.height = 90
        img.anchor = f'D{row}'
        ws.add_image(img)

    if len(ws._images) != len(items):
        raise RuntimeError(f'Image count {len(ws._images)} != product count {len(items)}')


def main():
    if not BACKUP.exists():
        shutil.copy2(SRC, BACKUP)
    shutil.copy2(BACKUP, SRC)

    wb = load_workbook(SRC)
    ws = wb.active
    merges = collect_merges(ws)
    unmerge_all(ws)

    # Roller merge (in place)
    ws.cell(9, 5, 'Detachable Manual Roller')
    ws.cell(
        9,
        6,
        'Standard or Narrow · Needle length: 0.25, 0.50, 1.00, 1.50, 2.00mm',
    )
    ws.cell(27, 5, 'Manual Roller')
    ws.cell(
        27,
        6,
        'Standard or Narrow · Needle length: 0.25, 0.50, 1.00, 1.50, 2.00mm',
    )
    for r in (11, 29):
        for c in range(5, 10):
            ws.cell(r, c).value = None

    # Delete narrow spacer rows bottom-up (unmerged, safe now)
    ws.delete_rows(29, 2)
    ws.delete_rows(11, 2)
    merges = shift_merges(merges, 29, -2)
    merges = shift_merges(merges, 11, -2)

    # Find peeling header after deletions
    peel_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if v and 'Peeling' in str(v) and 'Power' in str(v):
            peel_row = r
            break
    if not peel_row:
        raise RuntimeError('Peeling header not found')

    insert_at = peel_row
    insert_count = 5
    ws.insert_rows(insert_at, insert_count)
    merges = shift_merges(merges, insert_at, insert_count)

    # Bio Meso block
    ws.cell(insert_at, 4, 'GENOSYS Bio Meso PDRN')
    merges.append((insert_at, insert_at, 4, 9))

    r1, r2 = insert_at + 1, insert_at + 2
    merges.extend([(r1, r2, c, c) for c in range(4, 10)])
    ws.cell(r1, 5, 'BIO-MESO PDRN Expert Ampoule 60000')
    ws.cell(r1, 6, 'Professional bio-meso treatment ampoule')
    ws.cell(r1, 7, '3ml x 4 ampoules')
    ws.cell(r1, 8, 'Box')
    ws.cell(r1, 9, 300)

    r3, r4 = insert_at + 3, insert_at + 4
    merges.extend([(r3, r4, c, c) for c in range(4, 10)])
    ws.cell(r3, 5, 'BIO-MESO PDRN Homecare Ampoule 5000')
    ws.cell(r3, 6, 'Homecare maintenance ampoule')
    ws.cell(r3, 7, '50ml')
    ws.cell(r3, 8, 'Pcs')
    ws.cell(r3, 9, 150)

    apply_merges(ws, merges)

    # Style from mask template rows
    for c in range(1, 10):
        copy_style(ws.cell(33, c), ws.cell(insert_at, c))
    for c in range(1, 10):
        copy_style(ws.cell(38, c), ws.cell(r1, c))
        copy_style(ws.cell(38, c), ws.cell(r3, c))

    # Validate + CSV before image pass (uses same ws)
    items = extract_items(ws)
    if len(items) != 100:
        raise RuntimeError(f'Expected 100 priced lines, got {len(items)}')

    narrow = [x for x in items if 'Narrow' in x['product'] and 'Replacement' not in x['product']]
    if narrow:
        raise RuntimeError(f'Narrow rollers still present: {narrow}')

    bio = [x for x in items if x['category'] == 'GENOSYS Bio Meso PDRN']
    if len(bio) != 2:
        raise RuntimeError(f'Expected 2 Bio Meso lines, got {bio}')

    with CSV_OUT.open('w', newline='') as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                'row',
                'category',
                'product',
                'description',
                'quantity_or_spec',
                'unit',
                'price_aed',
            ],
        )
        w.writeheader()
        w.writerows(items)

    reanchor_images(ws, items)
    wb.save(SRC)

    print(f'OK: {len(items)} lines, {len(ws._images)} images → {SRC.name}')
    for x in bio:
        print(f"  {x['product']} @ {x['price_aed']} AED")


if __name__ == '__main__':
    main()
