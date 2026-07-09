#!/usr/bin/env python3
"""
Rebuild Korea export orderform — USD tab, column K (Order Qty.).

Data: MoySklad restock 2026-07-09 + season buffers (Jul 7 validated).

Usage: python3 scripts/populate-korea-export-orderform-20260709.py
"""

import os
import shutil
from datetime import datetime

import openpyxl

SRC = os.path.expanduser("~/Desktop/DTSMG_Export Orderform-2026_USD.xlsx")
STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
BACKUP = os.path.expanduser(
    f"~/Desktop/DTSMG_Export Orderform-2026_USD.before_korea_reorder_{STAMP}.xlsx"
)

# Korea item code -> order qty (form units: Pcs or Box per row)
ORDER_QTY = {
    # --- Tier A: CRITICAL (<30d cover) ---
    "GCMR02": 50,   # 54461 Makeup Remover — model 36
    "GCMA14": 250,  # 54467 PDRN mask pack — model 224, Q4 buffer
    "GCCR37": 230,  # 54457 Ultra Shield SPF50 — model 240
    "GCPS02": 36,   # 00020 SWS — 359 vials → 36 boxes ×10
    "GCPS05": 16,   # 00069 CTS — 153 vials → 16 boxes ×10
    # --- Tier B: URGENT (30–60d) ---
    "GCFO03": 110,  # 54464 Cushion Camel — model 107
    "GCCR44": 50,   # 00035 Problem Control Cream — model 51
    "GCSE17": 60,   # 00195 Hyaluron Serum — model 57
    "GCHR18": 26,   # 00048 HR³ Hair Solution box — model 26 (NEW vs Jul 8)
    # --- Tier C: season / Oct buffer / PLAN ---
    "GCMA10": 15,   # 00140 Sea Algae — consignment velocity (~1018/90d)
    "GCFO02": 100,  # 00144 Cushion Beige — model 98, Oct peak buffer
    "GCCR09": 40,   # 00041 Multi Sun SPF40 — summer stack, model 26
    "GCCR47": 40,   # 54473 Revita BB Natural — model 37
    "GCCR46": 20,   # 54472 Revita BB Bright — model 15
    "GCCR23": 15,   # 00037 Barrier Cream — model 12
    "GCCR07": 7,    # 00038 Post Cream 20g — 68 vials, clinic buffer (7 boxes ×12)
}


def main():
    if not os.path.isfile(SRC):
        raise SystemExit(f"Missing file: {SRC}")

    shutil.copy2(SRC, BACKUP)
    print(f"Backup: {BACKUP}")

    wb = openpyxl.load_workbook(SRC)
    ws = wb["USD"]

    cleared = 0
    for r in range(9, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or str(code).strip() == "":
            continue
        if ws.cell(r, 11).value not in (None, ""):
            ws.cell(r, 11).value = None
            cleared += 1

    set_rows = []
    missing = []
    for r in range(9, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code is None:
            continue
        code = str(code).strip()
        if code not in ORDER_QTY:
            continue
        qty = ORDER_QTY[code]
        ws.cell(r, 11).value = qty
        price = ws.cell(r, 10).value or 0
        try:
            amount = float(price) * qty
        except (TypeError, ValueError):
            amount = None
        set_rows.append((code, qty, price, amount))

    for code in ORDER_QTY:
        if code not in {c for c, *_ in set_rows}:
            missing.append(code)

    wb.save(SRC)

    print(f"Cleared {cleared} prior order-qty cells")
    print(f"Set {len(set_rows)} lines on: {SRC}\n")
    total_usd = 0.0
    for code, qty, price, amount in sorted(set_rows, key=lambda x: x[0]):
        if amount:
            total_usd += amount
        print(
            f"  {code}: qty {qty} @ ${price} = ${amount:.2f}"
            if amount
            else f"  {code}: qty {qty}"
        )
    print(f"\nEstimated subtotal: ${total_usd:,.2f} USD")
    if missing:
        print("WARNING — codes not found in sheet:", ", ".join(missing))


if __name__ == "__main__":
    main()
