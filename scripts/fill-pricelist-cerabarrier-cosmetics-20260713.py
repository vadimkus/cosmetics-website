#!/usr/bin/env python3
"""Fill CERABARRIER home + professional text into reserved GENOSYS Cosmetics rows."""
from __future__ import annotations

import csv
import io
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

XLSX = Path(
    '/Users/vadimkus/Desktop/Drive/Genosys/Price_list/2026/July/'
    'GENOSYS_UAE_PriceList_Clinics_2026.xlsx'
)
CSV_OUT = Path(__file__).resolve().parents[1] / 'docs/GENOSYS_UAE_PriceList_Clinics_2026_normalized.csv'
IMG_200 = Path(__file__).resolve().parents[1] / 'public/images/cera/cerabar_200ml.jpeg'
IMG_600 = Path(__file__).resolve().parents[1] / 'public/images/cera/cerabar_600ml.jpeg'

PRODUCT = 'CERABARRIER BIOME GEL CLEANSER'
ROWS = (
    (21, 'Daily biome gel cleanser / Personal use', '200ml', 190, IMG_200),
    (23, 'Professional biome gel cleanser', '600ml', 310, IMG_600),
)


def copy_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def add_image(ws, path: Path, row: int, col: int = 3):
    data = path.read_bytes()
    try:
        img = XLImage(io.BytesIO(data))
    except Exception:
        buf = io.BytesIO()
        Image.open(io.BytesIO(data)).save(buf, format='PNG')
        buf.seek(0)
        img = XLImage(buf)
    img.width = 90
    img.height = 90
    col_letter = chr(ord('A') + col - 1)
    img.anchor = f'{col_letter}{row}'
    ws.add_image(img)


def image_row(im) -> int | None:
    try:
        return im.anchor._from.row + 1
    except AttributeError:
        anchor = getattr(im, 'anchor', None)
        if isinstance(anchor, str):
            digits = ''.join(ch for ch in anchor if ch.isdigit())
            return int(digits) if digits else None
    return None


def remove_image_at_row(ws, row: int):
    kept = []
    for im in getattr(ws, '_images', []):
        if image_row(im) == row:
            continue
        kept.append(im)
    ws._images = kept


def extract_items(ws):
    items = []
    cat = None
    last_product = None
    for r in range(1, ws.max_row + 1):
        c3 = ws.cell(r, 3).value
        c4, c5, c6, c7, c8, c9 = [ws.cell(r, c).value for c in range(4, 10)]

        if c3 and str(c3).strip().startswith('GENOSYS') and not c4 and not c5:
            cat = str(c3).strip()
        if c4 and not c5 and not c6 and not c7 and not c8 and not c9:
            if 'GENOSYS' in str(c4):
                cat = str(c4).strip()

        price_raw = c9 if c9 is not None else c8
        if price_raw is None:
            continue
        pr = str(price_raw).strip()
        if pr in ('', 'Price                  AED'):
            continue
        if pr.upper() == 'N/A':
            price = 'N/A'
        else:
            try:
                price = float(price_raw)
            except (TypeError, ValueError):
                continue

        product = None
        description = ''
        qty = ''
        unit = ''
        if c5 and c4:
            product = str(c4).strip()
            last_product = product
            description = str(c5).strip()
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''
        elif c5:
            if c4:
                product = str(c4).strip()
                last_product = product
            else:
                product = last_product
            description = str(c5).strip()
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''
        elif c4:
            product = str(c4).strip()
            last_product = product
            description = str(c5).strip() if c5 else ''
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''

        if not product:
            continue
        last_product = product
        items.append(
            {
                'row': r,
                'category': cat or '',
                'product': product,
                'description': description,
                'quantity_or_spec': qty,
                'unit': unit,
                'price_aed': int(price) if isinstance(price, float) and price == int(price) else price,
            }
        )
    return items


def main():
    wb = load_workbook(XLSX)
    ws = wb.active
    template = 26

    for row, desc, spec, price, img_path in ROWS:
        for c in range(4, 9):
            copy_style(ws.cell(template, c), ws.cell(row, c))
        ws.cell(row, 4, PRODUCT)
        ws.cell(row, 5, desc)
        ws.cell(row, 6, spec)
        if not ws.cell(row, 7).value:
            ws.cell(row, 7, 'Pcs')
        ws.cell(row, 8, price)
        if img_path.exists():
            remove_image_at_row(ws, row)
            add_image(ws, img_path, row)

    wb.save(XLSX)

    items = extract_items(load_workbook(XLSX, data_only=True).active)
    cera = [x for x in items if 'CERABARRIER' in x['product'].upper()]
    if len(cera) != 2:
        raise RuntimeError(f'Expected 2 CERABARRIER lines, got {len(cera)}: {cera}')

    with CSV_OUT.open('w', newline='') as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                'row',
                'category',
                'product',
                'description',
                'quantity_or_spec',
                'unit',
                'price_aed',
            ],
        )
        w.writeheader()
        w.writerows(items)

    print(f'OK: {len(items)} priced lines')
    for x in cera:
        print(
            f"  R{x['row']}: {x['product']} | {x['description']} | "
            f"{x['quantity_or_spec']} | {x['unit']} | {x['price_aed']} AED"
        )


if __name__ == '__main__':
    main()
