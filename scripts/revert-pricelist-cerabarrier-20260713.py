#!/usr/bin/env python3
"""Revert CERABARRIER row insert from July 2026 clinic price list."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

XLSX = Path(
    '/Users/vadimkus/Desktop/Drive/Genosys/Price_list/2026/July/'
    'GENOSYS_UAE_PriceList_Clinics_2026.xlsx'
)
CSV_OUT = Path(__file__).resolve().parents[1] / 'docs/GENOSYS_UAE_PriceList_Clinics_2026_normalized.csv'

DELETE_AT = 73
DELETE_COUNT = 4


def collect_merges(ws):
    out = []
    for m in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(m))
        out.append((min_row, max_row, min_col, max_col))
    return out


def unmerge_all(ws):
    for m in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(m))
        except KeyError:
            pass


def apply_merges(ws, merges):
    seen = set()
    for min_row, max_row, min_col, max_col in merges:
        key = (min_row, max_row, min_col, max_col)
        if key in seen or min_row > max_row:
            continue
        seen.add(key)
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def rebuild_merges_after_delete(merges, delete_at: int, count: int):
    delete_end = delete_at + count - 1
    out = []
    for min_row, max_row, min_col, max_col in merges:
        if min_row >= delete_at and max_row <= delete_end:
            continue
        if min_row > delete_end:
            min_row -= count
            max_row -= count
        elif max_row < delete_at:
            pass
        else:
            continue
        out.append((min_row, max_row, min_col, max_col))
    return out


def shift_images_after_delete(ws, delete_at: int, count: int):
    kept = []
    delete_end = delete_at + count - 1
    for im in getattr(ws, '_images', []):
        row_1 = im.anchor._from.row + 1
        if delete_at <= row_1 <= delete_end:
            continue
        if row_1 >= delete_at + count:
            im.anchor._from.row -= count
            if hasattr(im.anchor, '_to') and im.anchor._to is not None:
                im.anchor._to.row -= count
        kept.append(im)
    ws._images = kept


def extract_items(ws):
    items = []
    cat = None
    last_product = None
    for r in range(1, ws.max_row + 1):
        c3 = ws.cell(r, 3).value
        c4, c5, c6, c7, c8, c9 = [ws.cell(r, c).value for c in range(4, 10)]

        if c3 and str(c3).strip().startswith('GENOSYS') and not c4 and not c5:
            cat = str(c3).strip()
        if c4 and not c5 and not c6 and not c7 and not c8 and not c9:
            if 'GENOSYS' in str(c4) or 'Genosys' in str(c4):
                cat = str(c4).strip()

        price_raw = c9 if c9 is not None else c8
        if price_raw is None:
            continue
        pr = str(price_raw).strip()
        if pr in ('', 'Price                  AED'):
            continue
        if pr.upper() == 'N/A':
            price = 'N/A'
        else:
            try:
                price = float(price_raw)
            except (TypeError, ValueError):
                continue

        product = None
        description = ''
        qty = ''
        unit = ''
        if c5 and c4:
            if c9 is not None and c8 is not None:
                product = str(c4).strip()
                last_product = product
                description = str(c5).strip()
                qty = str(c6).strip() if c6 else ''
                unit = str(c7).strip() if c7 else ''
            else:
                if c4:
                    product = str(c4).strip()
                    last_product = product
                else:
                    product = last_product
                description = str(c5).strip()
                qty = str(c6).strip() if c6 else ''
                unit = str(c7).strip() if c7 else ''
        elif c4:
            product = str(c4).strip()
            last_product = product
            description = str(c5).strip() if c5 else ''
            qty = str(c6).strip() if c6 else ''
            unit = str(c7).strip() if c7 else ''

        if not product:
            continue
        last_product = product
        items.append(
            {
                'row': r,
                'category': cat or '',
                'product': product,
                'description': description,
                'quantity_or_spec': qty,
                'unit': unit,
                'price_aed': int(price) if isinstance(price, float) and price == int(price) else price,
            }
        )
    return items


def main():
    wb = load_workbook(XLSX)
    ws = wb.active

    # sanity check
    if ws.cell(DELETE_AT, 4).value != 'CERABARRIER BIOME GEL CLEANSER':
        raise RuntimeError(
            f'Expected CERABARRIER at row {DELETE_AT}, got {ws.cell(DELETE_AT, 4).value!r}'
        )

    merges = rebuild_merges_after_delete(collect_merges(ws), DELETE_AT, DELETE_COUNT)
    unmerge_all(ws)
    shift_images_after_delete(ws, DELETE_AT, DELETE_COUNT)
    ws.delete_rows(DELETE_AT, DELETE_COUNT)
    apply_merges(ws, merges)
    wb.save(XLSX)

    wb2 = load_workbook(XLSX, data_only=True)
    items = extract_items(wb2.active)
    cera = [x for x in items if 'CERABARRIER' in x['product'].upper()]
    if cera:
        raise RuntimeError(f'CERABARRIER still present: {cera}')

    with CSV_OUT.open('w', newline='') as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                'row',
                'category',
                'product',
                'description',
                'quantity_or_spec',
                'unit',
                'price_aed',
            ],
        )
        w.writeheader()
        w.writerows(items)

    ws3 = load_workbook(XLSX).active
    toner_row = next(
        i
        for i in range(70, 85)
        if ws3.cell(i, 3).value and 'TONER' in str(ws3.cell(i, 3).value)
    )
    print(f'Reverted: {len(items)} priced lines, 0 CERABARRIER')
    print(f'TONER header now at R{toner_row}')


if __name__ == '__main__':
    main()
