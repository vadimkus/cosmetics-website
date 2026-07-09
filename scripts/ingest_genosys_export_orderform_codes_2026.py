#!/usr/bin/env python3
"""
Ingest the 2026 GENOSYS Export Order Form ("Codes.xlsx") into a normalized CSV.

Source: ~/Desktop/Exer/Codes.xlsx
Sheets: USD (products), GENOSYS Marketing Material, LED, HAIR-GENTRON

Output: docs/GENOSYS_Export_Orderform_Codes_2026_normalized.csv

This is the EXPORT order form (USD wholesale prices, item codes, barcodes,
labeling codes and HS customs codes) — distinct from the UAE clinic AED price list.

Usage: python3 scripts/ingest_genosys_export_orderform_codes_2026.py
"""
import csv
import os
import re
import openpyxl

SRC = os.path.expanduser("~/Desktop/Exer/Codes.xlsx")
OUT = os.path.join(os.path.dirname(__file__), "..", "docs",
                   "GENOSYS_Export_Orderform_Codes_2026_normalized.csv")

HEADER = ["sheet", "category", "itemCode", "barcode", "labeling", "hsCode",
          "description", "shortCode", "packSize", "unit", "priceUSD", "notes"]


def clean(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\n", " ")).strip()


def is_total(s):
    return bool(re.search(r"RANGE TOTAL|TOTAL AMOUNT", s, re.I))


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    rows = []

    # ---- USD sheet (main product catalogue) ----
    ws = wb["USD"]
    category = ""
    last_desc = ""
    for r in ws.iter_rows(values_only=True):
        r = list(r) + [None] * (15 - len(r))
        code = clean(r[0])
        barcode = clean(r[1])
        price = clean(r[9])
        if is_total(code):
            continue
        # Skip the repeated column-header rows
        if code in ("Products", "Item Code") or barcode == "Barcode":
            continue
        # Category / sub-heading header: first col only, no barcode, no price
        if code and not barcode and not price:
            category = code
            last_desc = ""
            continue
        if not code and not barcode:
            continue
        if not code:
            continue
        desc = clean(r[5]) or last_desc
        if clean(r[5]):
            last_desc = clean(r[5])
        rows.append({
            "sheet": "USD", "category": category, "itemCode": code,
            "barcode": barcode, "labeling": clean(r[2]), "hsCode": clean(r[3]),
            "description": desc, "shortCode": clean(r[6]), "packSize": clean(r[7]),
            "unit": clean(r[8]), "priceUSD": price, "notes": clean(r[12]),
        })

    # ---- Marketing Material sheet ----
    # cols: 2 Product Name | 3 Item Code | 4 Description | 5 Pack qty | 6 Material | 7 Unit | 8 Price USD
    ws = wb["GENOSYS Marketing Material"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    category = ""
    for r in grid[6:]:
        r = list(r) + [None] * (14 - len(r))
        section = clean(r[1])
        name = clean(r[2])
        code = clean(r[3])
        price = clean(r[8])
        if re.match(r"^Total$", name, re.I) or re.search(r"Goods Price fluctuate|Order for Starter", name, re.I):
            continue
        # Category header sits in col 1 (merged section title)
        if section and not name and not code:
            category = section
            continue
        if not code:
            continue
        desc = (name + " — " if name else "") + clean(r[4])
        rows.append({
            "sheet": "Marketing", "category": category, "itemCode": code,
            "barcode": "", "labeling": "", "hsCode": "", "description": desc,
            "shortCode": "", "packSize": clean(r[5]),
            "unit": clean(r[7]), "priceUSD": price, "notes": clean(r[6]),
        })

    # ---- LED sheet (single device, tiered pricing) ----
    rows.append({
        "sheet": "LED", "category": "GENO-LED IR", "itemCode": "GMPS05",
        "barcode": "", "labeling": "", "hsCode": "",
        "description": "GENO-LED IR II — 5 wavelengths (423/532/583/640/830nm); 1,145 LEDs; device+eyeshield+adaptor",
        "shortCode": "", "packSize": "1 device", "unit": "box",
        "priceUSD": "630 (1-4u) / 580 (5-9u) / 525 (10u+)", "notes": "Tiered PO pricing",
    })

    # ---- HAIR-GENTRON sheet ----
    ws = wb["HAIR-GENTRON"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    for r in grid[6:]:
        r = list(r) + [None] * (13 - len(r))
        code = clean(r[2])
        if not code or re.search(r"Total", code, re.I):
            continue
        rows.append({
            "sheet": "HAIR-GENTRON", "category": "HOUSEHOLD ELECTRICAL APPLIANCES",
            "itemCode": code, "barcode": "", "labeling": "", "hsCode": "",
            "description": clean(r[3]), "shortCode": "", "packSize": clean(r[4]),
            "unit": clean(r[5]), "priceUSD": clean(r[6]), "notes": "",
        })

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADER)
        w.writeheader()
        w.writerows(rows)

    by_sheet = {}
    for r in rows:
        by_sheet[r["sheet"]] = by_sheet.get(r["sheet"], 0) + 1
    print(f"Wrote {len(rows)} line items to {os.path.normpath(OUT)}")
    print("By sheet:", by_sheet)


if __name__ == "__main__":
    main()
