#!/usr/bin/env python3
"""Fill Genosys_SOA_CLEAN_NO_STAMP.xls with invoice rows and export PDF."""

import json
import subprocess
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

TEMPLATE_XLS = Path.home() / 'Desktop/Drive/Genosys/Print_forms/2026/invoice_ART/SOA/Genosys_SOA_CLEAN_NO_STAMP.xls'


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 12) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)


def ensure_xlsx(xls_path: Path, work_dir: Path) -> Path:
    xlsx_path = work_dir / 'soa_filled.xlsx'
    if not xls_path.exists():
        raise FileNotFoundError(f'Template not found: {xls_path}')
    subprocess.run(
        [
            'soffice',
            '--headless',
            '--convert-to',
            'xlsx',
            '--outdir',
            str(work_dir),
            str(xls_path),
        ],
        check=True,
        capture_output=True,
    )
    converted = work_dir / f'{xls_path.stem}.xlsx'
    if not converted.exists():
        raise RuntimeError('LibreOffice failed to convert SOA template to xlsx')
    converted.rename(xlsx_path)
    return xlsx_path


def unmerge_rows(ws, start_row: int, end_row: int) -> None:
    to_remove = []
    for mrange in list(ws.merged_cells.ranges):
        bounds = openpyxl.utils.range_boundaries(str(mrange))
        if bounds is None:
            continue
        min_col, min_row, max_col, max_row = bounds
        if max_row >= start_row and min_row <= end_row:
            to_remove.append(str(mrange))
    for mrange in to_remove:
        try:
            ws.unmerge_cells(mrange)
        except KeyError:
            pass


def fill_workbook(payload: dict, xlsx_path: Path) -> None:
    agent_name = payload['agentName']
    invoices = payload['invoices']
    statement_date = payload.get('statementDate') or datetime.now().strftime('%d/%m/%Y')
    total_paid = float(payload['totalPaid'])
    total_balance = float(payload['totalBalance'])

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    if ws['D2'].value and 'I14330AT' in str(ws['D2'].value):
        ws['D2'].value = str(ws['D2'].value).replace('I14330AT', '5023192')

    ws['C11'] = statement_date

    # Remove JXLS loop markers (delete bottom row first).
    ws.delete_rows(28, 1)
    ws.delete_rows(26, 1)

    data_start = 26
    n = len(invoices)
    if n > 1:
        ws.insert_rows(data_start + 1, n - 1)

    unmerge_rows(ws, data_start, data_start + n + 4)

    for i, inv in enumerate(invoices):
        row = data_start + i
        if i > 0:
            copy_row_style(ws, data_start, row)
        ws.cell(row, 1, inv['docNo'])
        ws.cell(row, 2, agent_name)
        ws.cell(row, 5, inv['date'])
        ws.cell(row, 6, float(inv['amount']))
        ws.cell(row, 7, float(inv['paid']))
        ws.cell(row, 8, float(inv['balance']))

    for row in range(data_start + n, data_start + n + 8):
        label = ws.cell(row, 5).value
        if label == 'Paid amount: AED':
            ws.cell(row, 8, total_paid)
        elif label == 'Pending payment: AED':
            ws.cell(row, 8, total_balance)

    wb.save(xlsx_path)


def xlsx_to_pdf(xlsx_path: Path, pdf_path: Path) -> None:
    subprocess.run(
        [
            'soffice',
            '--headless',
            '--convert-to',
            'pdf',
            '--outdir',
            str(pdf_path.parent),
            str(xlsx_path),
        ],
        check=True,
        capture_output=True,
    )
    generated = pdf_path.parent / f'{xlsx_path.stem}.pdf'
    if generated != pdf_path:
        if pdf_path.exists():
            pdf_path.unlink()
        generated.rename(pdf_path)


def main() -> None:
    payload = json.loads(Path(sys.argv[1]).read_text())
    out_pdf = Path(sys.argv[2])
    work_dir = out_pdf.parent / '.soa_work'
    work_dir.mkdir(parents=True, exist_ok=True)

    template = Path(payload.get('templatePath') or TEMPLATE_XLS)
    xlsx_path = ensure_xlsx(template, work_dir)
    fill_workbook(payload, xlsx_path)
    xlsx_to_pdf(xlsx_path, out_pdf)

    # Keep filled xlsx alongside PDF for audit.
    filled_xlsx = out_pdf.with_suffix('.xlsx')
    if filled_xlsx.exists():
        filled_xlsx.unlink()
    xlsx_path.rename(filled_xlsx)

    print(json.dumps({'pdf': str(out_pdf), 'xlsx': str(filled_xlsx)}))


if __name__ == '__main__':
    main()
